#!/public/home/zhangjie/anaconda3/envs/reinvent1/bin/python
import __main__
import pandas as pd
import argparse
from pathlib import Path
from rdkit import Chem, DataStructs
import os,sys
from rdkit.Chem import AllChem,Draw
from rdkit.Chem import MCS
from IPython.display import display, SVG,display_svg
from my_toolset.drawing_utils import *
from my_toolset.my_utils import get_mol
from pathlib import Path
from rdkit import RDLogger
import openpyxl
from openpyxl.drawing.image import Image 
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
import re,copy
lg = RDLogger.logger()
lg.setLevel(RDLogger.CRITICAL)

class argNameSpace():
    def __init__(self) -> None:
        self.A=""

def csvToExcel(csv, imgCols=['SMILES','smi'],save_file='',max_imgs=500,column=False):
    '''  Transform Pandas dataframe into Excel  '''
    df=pd.read_csv(csv)
    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'  ## the maximum columns cannot too long
    wb = openpyxl.Workbook()
    # ws = wb.create_sheet('Sheet1')
    ws = wb.active
    tmpImgPath=Path('./images_png')
    tmpImgPath.mkdir(exist_ok=True, parents=True)
    col_names=df.columns.to_list()
    ###  Write column head
    for icol,val in enumerate(col_names):
        irow=1
        if column:
            try:
            # if 1:  ## for dubugging
                col_letter=LETTERS[icol]
                ws.row_dimensions[irow].height = 90
                ws.column_dimensions[col_letter].width = 20
                mol = Chem.MolFromSmiles(val)
                img = Draw.MolToImage(mol, size=[200, 200])
                img.save(tmpImgPath.joinpath(f'molecule{irow}{icol}.png'))
                img=Image(tmpImgPath.joinpath(f'molecule{irow}{icol}.png'))
                img.width = 100
                img.height = 100
                # print(ws.cell(row=irow, column=icol).coordinate)
                img.anchor = ws.cell(row=irow, column=icol+1).coordinate # col and row are zero-based
                # ws.add_image(img, f"{col_letter}{irow}")
                ws.add_image(img)
            except Exception as e:
                print(e)
                continue
        ws.cell(irow,icol+1).value=val
    for idx,row in df.iterrows():
        irow=idx+2 ## the row id start from one and the first was occupied by head
        for icol,vcol in enumerate(col_names):
            if vcol in imgCols:
                if icol>max_imgs:
                    continue
                try:
                # if 1:  ## for dubugging
                    col_letter=LETTERS[icol]
                    ws.row_dimensions[irow].height = 90
                    ws.column_dimensions[col_letter].width = 25
                    mol = Chem.MolFromSmiles(row[vcol])
                    img = Draw.MolToImage(mol, size=[400, 200])
                    img.save(tmpImgPath.joinpath(f'molecule{irow}{icol}.png'))
                    img=Image(tmpImgPath.joinpath(f'molecule{irow}{icol}.png'))
                    img.width = 200
                    img.height = 100
                    # print(ws.cell(row=irow, column=icol).coordinate)
                    img.anchor = ws.cell(row=irow, column=icol+1).coordinate # col and row are zero-based
                    # ws.add_image(img, f"{col_letter}{irow}")
                    ws.add_image(img)
                except Exception as e:
                    print(e)
                    continue
            ws.cell(irow,icol+1).value=row[vcol]  ## add value to the folders
    wb.save(save_file)
    wb.close()
    os.system(f"rm -rf {tmpImgPath}")
    return wb

def main(args):
    df=pd.read_csv(args.csvFile)
    # df['Structure']=df[args.smilesCol]
    struct_file=args.csvFile.replace('.csv','')+'_struct.csv'
    df.to_csv(struct_file, index=None)
    excel_file=args.csvFile.replace('.csv','')+'.xlsx'
    csvToExcel(struct_file, imgCols=args.smilesCol,save_file=excel_file,max_imgs=1000,column=False)

def get_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument("--csvFile", help="the name of csv file", required=True, default='')
    parser.add_argument("--smilesCol", nargs='+', help="the column name of SMILES", required=False, default=[])
    args = parser.parse_args()
    return args

if __name__ == '__main__':
    test=0
    if test>0:
        args=argNameSpace
        args.maegzFiles=['Id_smi_dock_TLR7_pv_1p.maegz','Id_smi_dock_TLR7_pv_1p.maegz']
    else:
        args = get_parser()
    main(args)
